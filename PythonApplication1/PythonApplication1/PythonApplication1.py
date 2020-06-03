# ! /usr/bin/env python
# -*- coding: utf-8 -*-
import threading
import tkinter as tk
from tkinter import ttk
stack = []
osh=[]
labels=[]
all_sch={'all':0,'in':0,'good':0,'error':0}

def change_gal():
    text='all: {0}| in_prog: {1}| good: {2}| error: {3}'.format(all_sch['all'],all_sch['in'],all_sch['good'],all_sch['error'])
    print(text)
    inter.config(text=text)

def zap(p,indexxx):
    print('ind',indexxx)
    from selenium import webdriver
    from time import sleep
    from openpyxl import load_workbook

    wb = load_workbook(r'123.xlsx')
    for i in wb:
        sheet = wb[str(i)[12:-2]]
    kk = []
    while 1:
        if sheet["S" + str(p)].value == None: break
        brt = int(sheet["S" + str(p)].value)
        qq = []
        for ii in range(brt):
            qq.append(str(sheet["G" + str(p + ii)].value))
        pas_t = str(sheet["D" + str(p)].value)
        name = str(sheet["E" + str(p)].value)
        suname = str(sheet["AA" + str(p)].value)
        mail = str(sheet["AO" + str(p)].value)
        num_os = str(sheet["AD" + str(p)].value)
        if str(num_os)[0] != '0': num_os = '0' + num_os
        n = len(qq)


        GBs = []
        for ii in range(brt):
            GBs.append(str(sheet["V" + str(p + ii)].value))

        addr_town = str(sheet["AB" + str(p)].value)
        e = addr_town
        for ww in range(1, len(e)):
            if e[-ww] == ' ':
                addr_town = e[:-ww]
            else:
                break

        addrs = str(sheet["AC" + str(p)].value)
        for w in range(len(addrs)):
            try:
                int(addrs[w])
                addr_street = (addrs[:w - 1])
                e = addr_street
                for ww in range(1, len(e)):
                    if e[-ww] == ' ':
                        addr_street = e[:-ww]
                    else:
                        break
                break
            except:
                pass
        import re
        addr_numb_hous, addr_numb_kv = re.sub(r'\D', ' ', addrs).split()



        try:
            # logging
            kod = 4
            options = webdriver.ChromeOptions()

            driver2 = webdriver.Chrome(chrome_options=options, executable_path=r"webDriver/chromedriver.exe")

            url_osnovnoi = "https://www.hotmobile.co.il/retail/Pages/RetailPurchaseStep1.aspx"
            driver2.get(url_osnovnoi)
            sleep(10)

            element = driver2.find_element_by_id('RetailId').send_keys('elad1211')
            sleep(2)
            driver2.find_element_by_id('checkRetailer').click()
            sleep(3)
            driver2.find_element_by_id('AcStores').click()
            sleep(1)
            driver2.find_element_by_xpath('/html/body/div[1]/div[1]/div/div/div[2]/div/div/ul/li[1]').click()
            sleep(1)
            driver2.find_element_by_id('loginSubmit').click()
            sleep(2)

            print(n)
            kod=0
            for u in range(n):
                # 1

                numb = qq[u]
                GB = GBs[u]

                osnov = driver2.find_elements_by_class_name('formSection')[u]

                elements_control = osnov.find_element_by_class_name('controls')
                elements = elements_control.find_elements_by_tag_name('option')
                # print(len(elements))
                # for i in elements:#выбор gb
                #     print(i.get_attribute("innerHTML"))

                # gb

                if '80GB ב25 ש"ח לשנה' in GB:
                    for t in elements:
                        if 'מושלמת 50GB 2019 ב- 29 ₪ לחודש למשך שנה' in t.get_attribute("innerHTML"):t.click()

                elif '100GB ב29 ש"ח לשנה' in GB:
                    for t in elements:
                        if 'מושלמת 100GB בעלות 29.9 ₪ לחודש למשך שנה' in t.get_attribute("innerHTML"):t.click()

                elif '7GB ב19.9 ש"ח לשנה דמי חיבור 19.9' in GB:
                    for t in elements:
                        if 'מושלמת 7GB ב-19.9 ₪ לחודש למשך שנה' in t.get_attribute("innerHTML"):t.click()

                elif '150GB ב34.9 ש"ח -שלושה קווים ומעלה 5 שח הנחה מחיר לשנתיים' in GB:
                    for t in elements:
                        if 'מושלמת 150GB ב 34.9 ₪ לחודש למשך שנתיים (29.9 ש"ח במכירה של 3 מנויים ומעלה)' in t.get_attribute("innerHTML"):t.click()

                elif '80GB ב20 ש"ח לשנה' in GB:
                    for t in elements:
                        if 'מושלמת 50GB 2019 ב- 29 ₪ לחודש למשך שנה ' in t.get_attribute("innerHTML"):t.click()

                elif '100GB ב25.9 ש"ח לשנה מ2 קווים' in GB:
                    for t in elements:
                        if 'מושלמת 100GB בעלות 29.9 ₪ לחודש למשך שנה' in t.get_attribute("innerHTML"):t.click()

                elif '120GB ב22.9 ש"ח לשנה' in GB:
                    for t in elements:
                        if 'מושלמת 100GB בעלות 29.9 ₪ לחודש למשך שנה' in t.get_attribute("innerHTML"):t.click()

                elif '150GB חול כלול ב29.9 לשנתיים' in GB:
                    for t in elements:
                        if '51008494' in t.get_attribute("value"):t.click()

                elif '150GB פיקס ב29.9 ש"ח לשנתיים' in GB:
                    for t in elements:
                        if 'מושלמת 100GB עם NORTON+FIX ב-59 ₪ לחודש למשך שנתיים (כולל מע"מ)' in t.get_attribute("innerHTML"):t.click()

                elif '50GB ב15 ש"ח לשנה' in GB:
                    for t in elements:
                        if 'מושלמת 50GB ב14.9 ₪ לחודש למשך שנה' in t.get_attribute("innerHTML"):t.click()

                elif '150GB ב29.9 לכל החיים' in GB:
                    for t in elements:
                        if 'MAX 150GB ב 39.9 ש"ח לחודש (הנחת כמות 10 ש"ח במכירת 2 מנויים ומעלה)' in t.get_attribute("innerHTML"):t.click()

                elif '50GB ב24.9 לכל החיים' in GB:
                    for t in elements:
                        if 'MAX 50GB FOR LIFE ב 34.9 ש"ח לחודש (הנחת כמות 10 ש"ח במכירת 2 מנויים ומעלה)' in t.get_attribute("innerHTML"):t.click()


                # sim
                elements_sim = osnov.find_element_by_class_name('SimOptions')
                radio_sim = elements_sim.find_elements_by_class_name('originalListSim')
                # выбор sim

                if 1:  # первая
                    radio_sim[0].click()
                else:
                    radio_sim[1].click()

                # номер
                elements_sim_c = osnov.find_element_by_class_name('PortingContainer')
                radio_sim_o = elements_sim_c.find_elements_by_class_name('keepNumberRadio')

                if 'חדש' in numb:
                    radio_sim_o[1].find_element_by_tag_name('input').click()

                    sleep(1)
                    controls = osnov.find_elements_by_class_name('monthlyCostContainer')[3]
                    controls = controls.find_elements_by_tag_name('input')[1].click()
                    sleep(2)
                else:
                    radio_sim_o[0].find_element_by_tag_name('input').click()
                    if str(numb)[0] != '0': numb = '0' + numb
                    a = numb
                    b = a[:3]
                    list_numb = elements_sim_c.find_elements_by_tag_name('option')
                    if b == '050':
                        list_numb[1].click()
                    elif b == '051':
                        list_numb[2].click()
                    elif b == '052':
                        list_numb[3].click()
                    elif b == '053':
                        list_numb[4].click()
                    elif b == '054':
                        list_numb[5].click()
                    elif b == '055':
                        list_numb[6].click()
                    elif b == '058':
                        list_numb[7].click()
                    sleep(1)
                    elements_sim_c.find_element_by_class_name('formInput').send_keys(a[3:])
                    sleep(1)
                    elements_sim_c.find_element_by_class_name('btns_continueTitle35').click()
                    sleep(5)
                    # moreinfo

                    osnov = driver2.find_elements_by_class_name('formSection')[u]

                    try:
                        osnov.find_element_by_class_name('ui-dialog-buttonset').find_elements_by_tag_name(
                            'button').click()
                        sleep(1)
                    except:
                        pass

                    elements_sim_c = osnov.find_element_by_class_name('PortingContainer')
                    elements_pass = elements_sim_c.find_elements_by_class_name('keepNumberTR')

                    element_spis = elements_pass[4].find_elements_by_tag_name('option')[1].click()
                    sleep(1)
                    elements_pass[6].find_element_by_tag_name('input').send_keys(pas_t)
                    sleep(1)
                    elements_pass[8].find_element_by_tag_name('input').send_keys(name)
                    sleep(1)
                    elements_pass[10].find_element_by_tag_name('input').send_keys(suname)
                    sleep(1)
                    elements_pass[-1].find_element_by_tag_name('input').click()
                    sleep(1)

                moreinfo = osnov.find_element_by_class_name('moreinfo')
                selects = moreinfo.find_elements_by_tag_name('select')
                f = 1
                sleep(2)
                for i in selects[:-1]:
                    values = i.find_elements_by_tag_name('option')
                    if f == 1:
                        values[1].click()
                        f = 0
                    else:
                        values[2].click()
                    sleep(0.1)

                if 1:
                    tables = moreinfo.find_element_by_id('moreMblChk').find_elements_by_tag_name('table')
                    for i in tables:
                        r = i.find_element_by_tag_name('label').get_attribute("innerHTML")
                        if 'סימון לקוח חוזר' in r:
                            i.find_element_by_tag_name('input').click()
                            break
                if u + 1 == n:
                    try:
                        sleep(3)
                        driver2.find_element_by_xpath('//*[@id="subscriberForm"]/div[7]/a/div[2]/span').click()
                    except:
                        sleep(3)
                        driver2.find_element_by_class_name('continue').find_element_by_tag_name('a').click()
                else:
                    try:
                        sleep(1)
                        driver2.find_element_by_xpath('//*[@id="subscriberForm"]/div[8]/a/div[2]/span').click()
                    except:
                        sleep(2)
                        driver2.find_element_by_class_name('add').find_element_by_tag_name('a').click()
                        ##.find_element_by_tag_name('a')
                    sleep(7)


            kod = 1

            # 2
            sector = driver2.find_element_by_class_name('formSection')
            sector.find_element_by_class_name('newCustomer').find_element_by_tag_name('input').click()
            sleep(1)
            sector.find_element_by_class_name('mishluach').find_elements_by_tag_name('label')[1].click()

            sector.find_element_by_class_name('txbEmail').send_keys(mail)

            controls = sector.find_element_by_class_name('controls').find_elements_by_tag_name('tr')
            controls[0].find_element_by_tag_name('input').send_keys(name)
            controls[1].find_element_by_tag_name('input').send_keys(suname)
            controls[2].find_element_by_tag_name('input').send_keys(pas_t)

            for ii in range(2):
                aa = num_os[:3]

                controls[3 + ii].find_element_by_tag_name('input').send_keys(num_os[3:])

                for iii in controls[3 + ii].find_elements_by_tag_name('option'):
                    if aa in iii.get_attribute("innerHTML"):
                        iii.click()
                        break

            controls[5].find_element_by_tag_name('input').send_keys(mail)

            pod = 0
            tttt = [7]
            for ii in range(1):  # town
                print('tow')
                tt = tttt[ii]
                for i in range(len(addr_town)):
                    controls[tt].find_element_by_tag_name('input').send_keys(addr_town[i])
                    sleep(1.5)
                    # try:
                    lis = driver2.find_elements_by_class_name('ac_results')
                    if len(lis) > 0:
                        lis = driver2.find_elements_by_class_name('ac_results')[0].find_elements_by_tag_name('li')
                        pod_t = len(lis)

                        if pod_t == 1:
                            lis[0].click()
                            break
                        elif pod_t == 0 and pod != 0:
                            controls[tt].find_element_by_tag_name('input').clear()
                            controls[tt].find_element_by_tag_name('input').send_keys(addr_town[:i])
                            lis = driver2.find_elements_by_class_name('ac_results')[0].find_elements_by_tag_name('li')[
                                0].click()
                            break
                        elif i + 1 == len(addr_town):
                            lis[0].click()
                            sleep(2)
                            break
                    # except:pass
            tttt = [8]
            pod = 0
            for ii in range(1):  # town
                print('street')
                tt = tttt[ii]
                for i in range(len(addr_street)):
                    controls[tt].find_element_by_tag_name('input').send_keys(addr_street[i])
                    sleep(2)
                    # try:
                    lis = driver2.find_elements_by_class_name('ac_results')
                    if len(lis) > 1:
                        lis = driver2.find_elements_by_class_name('ac_results')[1].find_elements_by_tag_name('li')
                        pod_t = len(lis)

                        if pod_t == 1:
                            lis[0].click()
                            break
                        elif pod_t == 0 and pod != 0:
                            controls[tt].find_element_by_tag_name('input').clear()
                            controls[tt].find_element_by_tag_name('input').send_keys(addr_street[:i])
                            lis = driver2.find_elements_by_class_name('ac_results')[1].find_elements_by_tag_name('li')[
                                0].click()
                            break
                        elif i + 1 == len(addr_street):
                            lis[0].click()
                            sleep(2)

            sleep(0.1)
            controls[10].find_element_by_tag_name('input').send_keys(addr_numb_hous)
            sleep(0.1)
            controls[11].find_element_by_tag_name('input').send_keys(addr_numb_kv)
            sleep(0.1)
            controls[9].find_element_by_tag_name('input').send_keys('000')
            sleep(0.1)

            second = sector.find_element_by_xpath(
                '/html/body/form/div[7]/div/div[3]/div/div[3]/div[2]/div[1]/div[2]/table/tbody/tr/td/table/tbody/tr/td/div/div/div[2]/div[1]/div[8]/div[2]/div').find_elements_by_tag_name(
                'tr')

            pod = 0
            tttt = [1]
            for ii in range(1):  # town
                tt = tttt[ii]
                for i in range(len(addr_town)):
                    second[tt].find_element_by_tag_name('input').send_keys(addr_town[i])
                    sleep(1.5)
                    # try:
                    lis = driver2.find_elements_by_class_name('ac_results')
                    if len(lis) > 2:
                        lis = driver2.find_elements_by_class_name('ac_results')[2].find_elements_by_tag_name('li')
                        pod_t = len(lis)

                        if pod_t == 1:
                            lis[0].click()
                            break
                        elif pod_t == 0 and pod != 0:
                            second[tt].find_element_by_tag_name('input').clear()
                            second[tt].find_element_by_tag_name('input').send_keys(addr_town[:i])
                            lis = driver2.find_elements_by_class_name('ac_results')[2].find_elements_by_tag_name('li')[
                                0].click()
                            break
                        elif i + 1 == len(addr_town):
                            lis[0].click()
                            sleep(2)
                            break
            tttt = [3]
            pod = 0
            for ii in range(1):  # town
                tt = tttt[ii]
                for i in range(len(addr_street)):
                    second[tt].find_element_by_tag_name('input').send_keys(addr_street[i])
                    sleep(2)
                    # try:
                    lis = driver2.find_elements_by_class_name('ac_results')
                    if len(lis) > 3:
                        lis = driver2.find_elements_by_class_name('ac_results')[3].find_elements_by_tag_name('li')
                        pod_t = len(lis)

                        if pod_t == 1:
                            lis[0].click()
                            break
                        elif pod_t == 0 and pod != 0:
                            second[tt].find_element_by_tag_name('input').clear()
                            second[tt].find_element_by_tag_name('input').send_keys(addr_street[:i])
                            lis = driver2.find_elements_by_class_name('ac_results')[3].find_elements_by_tag_name('li')[
                                0].click()
                            break
                        elif i + 1 == len(addr_street):
                            lis[0].click()
                            sleep(2)
                            break

            second[4].find_element_by_tag_name('input').send_keys(addr_numb_hous)
            sleep(0.1)
            second[5].find_element_by_tag_name('input').send_keys(addr_numb_kv)

            while osh[indexxx]==0:
                sleep(1)

            driver2.find_element_by_class_name('exist').find_element_by_tag_name('a').click()
            sleep(5)



            kod = 2

            # 3
            osn = driver2.find_element_by_class_name('controls')

            nuz = str(sheet["AG" + str(p)].value)

            if 'הוראת קבע' not in nuz:

                card = str(sheet["AP" + str(p)].value)
                g = str(sheet["AQ" + str(p)].value)
                print(len(g), g)
                if g != 'None':
                    g_1, g_2 = g.split('/')

                cards = osn.find_elements_by_class_name('phCreditCardPayment')
                sleep(0.5)
                opt = cards[0].find_elements_by_tag_name('option')[1].click()
                sleep(0.5)
                cards[1].find_element_by_tag_name('input').send_keys(card)
                sleep(0.5)
                cards[2].find_element_by_tag_name('input').send_keys(pas_t)
                gods = cards[3].find_elements_by_tag_name('select')
                for i in gods[0].find_elements_by_tag_name('option'):
                    if g_1 in i.get_attribute("innerHTML"):
                        i.click()
                        break
                sleep(1)
                for i in gods[1].find_elements_by_tag_name('option'):
                    if g_2 in i.get_attribute("innerHTML"):
                        i.click()
                        break

                sleep(1)

            else:
                per = driver2.find_element_by_class_name('radioInp').find_elements_by_tag_name('input')[1].click()
                numb = driver2.find_element_by_class_name('bankchargeFormsInner').find_elements_by_tag_name('input')[
                    1].send_keys(num_os)
                bank = str(sheet["AS" + str(p)].value)

                bb = driver2.find_elements_by_class_name('phBankOrderPayment')[0].find_elements_by_tag_name('option')
                for iiii in bb:
                    if bank in iiii.get_attribute("innerHTML"): iiii.click()

                bank_numb = str(sheet["AT" + str(p)].value)
                numb = driver2.find_elements_by_class_name('phBankOrderPayment')[1].find_elements_by_tag_name('input')[
                    0].send_keys(bank_numb)
                bank_numb2 = str(sheet["AU" + str(p)].value)
                numb = driver2.find_elements_by_class_name('phBankOrderPayment')[2].find_elements_by_tag_name('input')[
                    0].send_keys(bank_numb2)

                numb = driver2.find_elements_by_class_name('phBankOrderPayment')[3].find_elements_by_tag_name('input')[
                    0].send_keys(pas_t)

            driver2.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            sleep(1)
            div = driver2.find_element_by_class_name('LBD_CaptchaDiv').find_element_by_tag_name('a').screenshot(
                "screenshot_с.png")

            import base64
            from python_rucaptcha import ImageCaptcha

            RUCAPTCHA_KEY = "e6f3776bacfcd04e3b47db309ebe4ac4"
            with open('screenshot_с.png', "rb") as image_file:
                encoded_string = base64.b64encode(image_file.read())
                user_answer = ImageCaptcha.ImageCaptcha(rucaptcha_key=RUCAPTCHA_KEY).captcha_handler(
                    captcha_base64=encoded_string)

            if not user_answer['error']:
                captch = (user_answer['captchaSolve'])
            elif user_answer['error']:
                print(user_answer['errorBody']['text'])

            driver2.find_element_by_class_name('formInput').send_keys(captch)
            sleep(1)
            driver2.find_element_by_class_name('payAggreementTR').find_element_by_tag_name('input').click()
            btn_fin = driver2.find_element_by_class_name('CreditCardFinisBtn').find_element_by_tag_name('a').click()
            sleep(10)

            if 'https://www.hotmobile.co.il/retail/pages/retailpurchasefinish.aspx' in driver2.current_url:
                ord = driver2.find_element_by_class_name('order')
                kod1 = ord.find_element_by_tag_name('h4').find_element_by_tag_name('span').get_attribute("innerHTML")
                kod2 = ord.find_element_by_class_name('spaceerMar').get_attribute("innerHTML")

                sheet["AN" + str(p)].value = kod1 + ' ' + kod2
                labels[indexxx].config(text=kod1 + ' ' + kod2)

                all_sch['in'] -= 1
                all_sch['good'] += 1
                change_gal()

                p += brt
                driver2.close()
                wb.save('123.xlsx')
                sleep(2)
            else:
                try:

                    all_sch['in'] -= 1
                    all_sch['error'] += 1
                    change_gal()

                    sheet["AN" + str(p)].value = driver2.find_element_by_class_name('wraper').get_attribute("innerHTML")
                    labels[indexxx].config(text=driver2.find_element_by_class_name('wraper').get_attribute("innerHTML"))
                    p += brt
                    wb.save('123.xlsx')
                    sleep(2)
                    kk.append(driver2)

                except:
                    break
            p += brt
            wb.save('123.xlsx')
            kk.append(driver2)
            sleep(2)
        except:
            text=''
            if kod == 0:
                text='error on 1'

            elif kod == 1:
                text='error on 2'
            elif kod == 2:
                text='error on 3'
            elif kod == 4:
                text='error on log'

            all_sch['in']-=1
            all_sch['error']+=1
            change_gal()

            err = []
            erorrs = driver2.find_elements_by_class_name('errMsg')
            for ii in erorrs:
                t = ii.find_elements_by_tag_name('span')
                for iii in t:
                    if 'inline' in iii.get_attribute('style'):
                        err.append(iii.get_attribute("innerHTML"))
            print(len(err))
            text+=' '+'|'.join(err)
            sheet["AN" + str(p)].value = text

            labels[indexxx].config(text=text)

            p += brt
            wb.save('123.xlsx')
            kk.append(driver2)
        input()

def o12():
    print('o12 start')
    global osh,stack,labels,all_sch,inter

    class ScrollableFrame(ttk.Frame):
        def __init__(self, container, *args, **kwargs):
            super().__init__(container, *args, **kwargs)
            canvas = tk.Canvas(self, width=1500, height=600)
            scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
            self.scrollable_frame = ttk.Frame(canvas)

            self.scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(
                    scrollregion=canvas.bbox("all")
                )
            )

            canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

            canvas.configure(yscrollcommand=scrollbar.set)

            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

    root = tk.Tk()

    root.resizable(width=False, height=False)
    canvas = tk.Canvas(root, background="grey90")
    canvas.pack(side="bottom", fill="both", expand=True)
    frame = ScrollableFrame(canvas, width=1000, height=600)

    def delit(a, b):
        all_sch['in']+=1
        change_gal()
        osh[b] = 0
        stack.append([a, b])

    def oshr(a):
        osh[a] = 1

    d_all=[]
    def delit_all():
        print(d_all)
        for i in d_all:
            all_sch['in'] += 1
            change_gal()
            stack.append(i)

    def osh_all():
        for i in range(len(osh)):
            osh[i]=1


    def change(ind):

        def save(ind, ms_name, sr_name, nums, schts):
            brt = int(sheet["S" + str(ind)].value)
            for i in range(brt):
                # sheet["E" + str(ind + i)].value = ms_name.get()
                # sheet["AA" + str(ind + i)].value = sr_name.get()
                sheet["G" + str(ind + i)].value = nums[i].get()
                sheet["V" + str(ind + i)].value = schts[i].get()

            wb.save('123.xlsx')

        root1 = tk.Tk()
        root1.resizable(width=False, height=False)

        brt = int(sheet["S" + str(ind)].value)
        canvas2 = tk.Canvas(root1, background="grey90", width=500, height=60 + 40 * brt)
        canvas2.pack(side="bottom", fill="both", expand=True)

        # name = str(sheet["E" + str(ind)].value)
        # suname = str(sheet["AA" + str(ind)].value)
        #
        # messages = tk.StringVar()
        #
        # message_entry1 = tk.Entry(canvas2, textvariable=messages)
        # message_entry1.place(x=5, y=5, width=100)
        # message_entry1.insert(0, name)
        # messages2 = tk.StringVar()
        # message_entry2 = tk.Entry(canvas2, textvariable=messages2)
        # message_entry2.place(x=110, y=5, width=100)
        # message_entry2.insert(0, suname)

        lis_v = ['80GB ב25 ש"ח לשנה', '100GB ב29 ש"ח לשנה', '7GB ב19.9 ש"ח לשנה דמי חיבור 19.9',
                 '150GB ב34.9 ש"ח -שלושה קווים ומעלה 5 שח הנחה מחיר לשנתיים', '80GB ב20 ש"ח לשנה',
                 '100GB ב25.9 ש"ח לשנה מ2 קווים', '120GB ב22.9 ש"ח לשנה', '150GB חול כלול ב29.9 לשנתיים',
                 '150GB פיקס ב29.9 ש"ח לשנתיים','50GB ב15 ש"ח לשנה','150GB ב29.9 לכל החיים','50GB ב24.9 לכל החיים']

        numss = []
        kreds = []
        for ii in range(brt):
            nums = (str(sheet["G" + str(ind + ii)].value))
            tar = (str(sheet["V" + str(ind + ii)].value))

            mes = tk.StringVar()
            message_entry = tk.Entry(canvas2, textvariable=mes)
            message_entry.place(x=5, y=30 + 40 * ii, width=205)
            message_entry.insert(0, nums)

            numss.append(message_entry)

            var = tk.StringVar(root1)
            var.set(tar)
            opt = tk.OptionMenu(canvas2, var, *lis_v)
            opt.place(x=210, y=30 + 40 * ii, width=205)
            kreds.append(var)
        message_button = tk.Button(canvas2, text="Save",
                                   command=lambda a=ind, nm=0, sm=0, nms=numss, k=kreds: save(
                                       a, nm, sm, nms, k))
        message_button.place(x=210, y=30 + 40 * (ii + 1))
        root1.mainloop()

    def INFO(xy):
        root1 = tk.Tk()
        root1.resizable(width=False, height=False)
        canvas2 = tk.Canvas(root1, background="grey90", width=420, height=80)
        canvas2.pack(side="bottom", fill="both", expand=True)
        message_entry = tk.Label(canvas2)
        message_entry.place(x=10, y=10, width=400, height=60)
        message_entry.config(text=str(sheet[xy].value))

        root1.mainloop()

    def change_one(y,ind,sk,lab):

        def save_one(y,ind,sk,lab,en):
            for i in range(sk):
                sheet[y + str(ind + i)].value = en.get()
            lab.config(text=en.get())
            wb.save('123.xlsx')

        root1 = tk.Tk()
        root1.resizable(width=False, height=False)
        canvas2 = tk.Canvas(root1, background="grey90", width=420, height=80)
        canvas2.pack(side="bottom", fill="both", expand=True)
        message_entry = tk.Entry(canvas2)
        message_entry.place(x=10, y=10, width=400, height=20)

        z=str(sheet[y + str(ind)].value)

        message_entry.insert(0, z)

        btn4 = tk.Button(canvas2, text='Save', command=lambda a=y,b=ind,c=sk,d=lab,y=message_entry: save_one(a,b,c,d,y))
        btn4.place(x=10, y=40)

        root1.mainloop()



    from openpyxl import load_workbook

    wb = load_workbook(r'123.xlsx')

    for i in wb:
        sheet = wb[str(i)[12:-2]]

    p = 2
    canv = tk.Frame(frame.scrollable_frame, width=10, height=10, background="grey90")
    canv.pack(fill='x', expand=False)

    lab = ttk.Label(canv, text=str('Line'), width=5)
    lab.pack(side='left')

    lab = ttk.Label(canv, text='Name', width=10)
    lab.pack(side='left')
    lab = ttk.Label(canv, text='Surname', width=10)
    lab.pack(side='left')
    lab = ttk.Label(canv, text='Passport', width=20)
    lab.pack(side='left')
    lab = ttk.Label(canv, text='Town', width=15)
    lab.pack(side='left')
    lab = ttk.Label(canv, text='Adress', width=20)
    lab.pack(side='left')
    lab = ttk.Label(canv, text='Pay', width=20)
    lab.pack(side='left')

    inter = ttk.Label(canv, text='Pay', width=40)
    inter.pack(side='left')

    btn4 = ttk.Button(canv, text="Начать Все", command=delit_all)
    btn4.pack(side='left')

    btn4 = ttk.Button(canv, text="Продолжить Все", command=osh_all)
    btn4.pack(side='left')

    iii = 0
    print('sev')
    while 1:
        if sheet["S" + str(p)].value == None: break
        brt = int(sheet["S" + str(p)].value)
        qq = []
        for ii in range(brt):
            qq.append(str(sheet["G" + str(p + ii)].value))
        pas_t = str(sheet["D" + str(p)].value)
        name = str(sheet["E" + str(p)].value)
        suname = str(sheet["AA" + str(p)].value)
        mail = str(sheet["AO" + str(p)].value)
        us = str(sheet["V" + str(p + ii)].value)

        addr_town = str(sheet["AB" + str(p)].value)

        addrs = str(sheet["AC" + str(p)].value)

        nuz = str(sheet["AG" + str(p)].value)

        canv = tk.Frame(frame.scrollable_frame, width=10, height=10, background="grey90")
        canv.pack(fill='x', expand=False)

        btn4 = ttk.Button(canv, text=str(brt), command=lambda a=p: change(a), width=5)
        btn4.pack(side='left')

        lab = ttk.Label(canv, text=str(name), width=8)
        lab.pack(side='left')
        btn4 = ttk.Button(canv, text='<', command=lambda a='E',ind=p,sk=brt,l=lab: change_one(a,ind,sk,l), width=2)
        btn4.pack(side='left')

        lab = ttk.Label(canv, text=str(suname), width=8)
        lab.pack(side='left')
        btn4 = ttk.Button(canv, text='<', command=lambda a='AA',ind=p,sk=brt,l=lab: change_one(a,ind,sk,l), width=2)
        btn4.pack(side='left')

        lab = ttk.Label(canv, text=str(pas_t), width=18)
        lab.pack(side='left')
        btn4 = ttk.Button(canv, text='<', command=lambda a='D',ind=p,sk=brt,l=lab: change_one(a,ind,sk,l), width=2)
        btn4.pack(side='left')

        lab = ttk.Label(canv, text=addr_town, width=13)
        lab.pack(side='left')
        btn4 = ttk.Button(canv, text='<', command=lambda a='AB',ind=p,sk=brt,l=lab: change_one(a,ind,sk,l), width=2)
        btn4.pack(side='left')

        lab = ttk.Label(canv, text=addrs, width=18)
        lab.pack(side='left')
        btn4 = ttk.Button(canv, text='<', command=lambda a='AC',ind=p,sk=brt,l=lab: change_one(a,ind,sk,l), width=2)
        btn4.pack(side='left')


        lab = ttk.Label(canv, text=nuz, width=18)
        lab.pack(side='left')
        btn4 = ttk.Button(canv, text='<', command=lambda a='AG',ind=p,sk=brt,l=lab: change_one(a,ind,sk,l), width=2)
        btn4.pack(side='left')

        if 'הוראת קבע' not in nuz:
            card = str(sheet["AP" + str(p)].value)
            g = str(sheet["AQ" + str(p)].value)
            lab = ttk.Label(canv, text=card, width=13)
            lab.pack(side='left')
            btn4 = ttk.Button(canv, text='<', command=lambda a='AP', ind=p, sk=brt, l=lab: change_one(a, ind, sk, l),
                              width=2)
            btn4.pack(side='left')

            lab = ttk.Label(canv, text=g, width=13)
            lab.pack(side='left')
            btn4 = ttk.Button(canv, text='<', command=lambda a='AQ', ind=p, sk=brt, l=lab: change_one(a, ind, sk, l),
                              width=2)
            btn4.pack(side='left')

            lab = ttk.Label(canv, text='', width=13)
            lab.pack(side='left')

            btn4 = ttk.Button(canv, text='',
                              width=2)
            btn4.pack(side='left')
        else:
            bank = str(sheet["AS" + str(p)].value)
            bank_numb = str(sheet["AT" + str(p)].value)
            bank_numb2 = str(sheet["AU" + str(p)].value)

            lab = ttk.Label(canv, text=bank, width=13)
            lab.pack(side='left')
            btn4 = ttk.Button(canv, text='<', command=lambda a='AS', ind=p, sk=brt, l=lab: change_one(a, ind, sk, l),
                              width=2)
            btn4.pack(side='left')

            lab = ttk.Label(canv, text=bank_numb, width=13)
            lab.pack(side='left')
            btn4 = ttk.Button(canv, text='<', command=lambda a='AT', ind=p, sk=brt, l=lab: change_one(a, ind, sk, l),
                              width=2)
            btn4.pack(side='left')

            lab = ttk.Label(canv, text=bank_numb2, width=13)
            lab.pack(side='left')
            btn4 = ttk.Button(canv, text='<', command=lambda a='AU', ind=p, sk=brt, l=lab: change_one(a, ind, sk, l),
                              width=2)
            btn4.pack(side='left')

        lab = ttk.Label(canv, text='', width=30)
        lab.pack(side='left')
        labels.append(lab)

        btn4 = ttk.Button(canv, text="I", command=lambda a="AN" + str(p): INFO(a), width=2)
        btn4.pack(side='left')
        d_all.append([p,iii])
        btn4 = ttk.Button(canv, text="Начать", command=lambda a=p, b=iii: delit(a, b))
        btn4.pack(side='left')

        btn4 = ttk.Button(canv, text="Продолжить", command=lambda a=iii: oshr(a))
        btn4.pack(side='left')
        osh.append(0)
        p += brt
        print(p)
        iii += 1
        all_sch['all']+=1

    change_gal()

    frame.pack()
    root.mainloop()

p12 = threading.Thread(target=o12, name="t61", args=[])
p12.start()

def o3():
    import time
    time.sleep(5)
    print('start')
    while 1:
        try:
            if len(stack)>0:
                print('za')
                i=stack.pop(0)
                print(i)

                def o4():
                    zap(i[0], i[1])

                p4 = threading.Thread(target=o4, args=[])
                p4.start()


            time.sleep(1)
        except:pass

p3 = threading.Thread(target=o3, args=[])
p3.start()


